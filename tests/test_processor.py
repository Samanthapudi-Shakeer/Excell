from __future__ import annotations

import io
import xml.etree.ElementTree as ET
import zipfile

from openpyxl import Workbook, load_workbook
from openpyxl.chart import BarChart, Reference
from openpyxl.comments import Comment
from openpyxl.styles import Font

from excel_translator.processor import process_excel_file


def _sample_workbook_bytes() -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Sales"
    ws["A1"] = "Hello"
    ws["A2"] = 10
    ws["B2"] = 20
    ws["C2"] = "=A2+B2"
    ws["A3"] = "Merged text"
    ws.merge_cells("A3:B3")
    ws["A1"].font = Font(bold=True, color="00FF0000")
    ws["A1"].comment = Comment("Review this", "qa")

    chart = BarChart()
    chart.title = "Quarterly Revenue"
    chart.y_axis.title = "Amount"
    chart.x_axis.title = "Quarter"
    data = Reference(ws, min_col=1, min_row=2, max_col=2, max_row=2)
    chart.add_data(data, titles_from_data=False)
    ws.add_chart(chart, "E2")

    ws2 = wb.create_sheet("Ops")
    ws2["A1"] = "World"

    buf = io.BytesIO()
    wb.save(buf)
    wb.close()
    return buf.getvalue()


def _inject_custom_drawing(workbook_bytes: bytes) -> bytes:
    drawing_xml = b'''<?xml version="1.0" encoding="UTF-8"?>
<xdr:wsDr xmlns:xdr="http://schemas.openxmlformats.org/drawingml/2006/spreadsheetDrawing" xmlns:a="http://schemas.openxmlformats.org/drawingml/2006/main">
  <xdr:twoCellAnchor>
    <xdr:sp>
      <xdr:txBody>
        <a:p><a:r><a:t>Flowchart Step</a:t></a:r></a:p>
      </xdr:txBody>
    </xdr:sp>
  </xdr:twoCellAnchor>
</xdr:wsDr>'''

    in_buf = io.BytesIO(workbook_bytes)
    out_buf = io.BytesIO()
    with zipfile.ZipFile(in_buf, "r") as zin, zipfile.ZipFile(out_buf, "w", zipfile.ZIP_DEFLATED) as zout:
        for info in zin.infolist():
            zout.writestr(info, zin.read(info.filename))
        zout.writestr("xl/drawings/drawing99.xml", drawing_xml)

    return out_buf.getvalue()


def _chart_a_t_texts(xlsx_bytes: bytes) -> list[str]:
    with zipfile.ZipFile(io.BytesIO(xlsx_bytes), "r") as zf:
        chart_parts = sorted(name for name in zf.namelist() if name.startswith("xl/charts/chart") and name.endswith(".xml"))
        assert chart_parts
        root = ET.fromstring(zf.read(chart_parts[0]))
        return [node.text for node in root.iter("{http://schemas.openxmlformats.org/drawingml/2006/main}t") if node.text]


def _drawing_texts(xlsx_bytes: bytes, path: str) -> list[str]:
    with zipfile.ZipFile(io.BytesIO(xlsx_bytes), "r") as zf:
        assert path in zf.namelist()
        root = ET.fromstring(zf.read(path))
        return [node.text for node in root.iter("{http://schemas.openxmlformats.org/drawingml/2006/main}t") if node.text]


def test_translation_preserves_formula_and_formatting(monkeypatch):
    from excel_translator import processor

    monkeypatch.setattr(
        processor.RoutedTranslator,
        "translate_with_engine",
        lambda self, text, source, target: (f"T[{text}]", "fake_engine"),
    )

    result = process_excel_file(
        file_name="input.xlsx",
        file_bytes=_inject_custom_drawing(_sample_workbook_bytes()),
        source_lang="en",
        target_lang="fr",
        selected_engine="azure",
    )

    wb = load_workbook(io.BytesIO(result.output_bytes))
    ws = wb[wb.sheetnames[0]]

    assert ws["A1"].value == "T[Hello]"
    assert ws["C2"].value == "=A2+B2"
    assert ws["A1"].font.bold is True
    assert ws["A1"].font.color.rgb == "00FF0000"
    assert "A3:B3" in [str(rng) for rng in ws.merged_cells.ranges]
    assert ws["A1"].comment.text == "T[Review this]"

    chart_texts = _chart_a_t_texts(result.output_bytes)
    assert "T[Quarterly Revenue]" in chart_texts
    assert "T[Amount]" in chart_texts
    assert "T[Quarter]" in chart_texts

    drawing_texts = _drawing_texts(result.output_bytes, "xl/drawings/drawing99.xml")
    assert drawing_texts == ["T[Flowchart Step]"]

    assert len(wb.sheetnames) == 2
    assert result.output_filename == "input_fr.xlsx"
    assert any(log.object_id == "sheet_title" for log in result.logs)

    wb.close()
