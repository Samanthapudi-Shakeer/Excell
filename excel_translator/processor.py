from __future__ import annotations

import io
import re
from dataclasses import dataclass
from pathlib import Path
from typing import List

from openpyxl import load_workbook
from openpyxl.cell.cell import Cell

from .drawing_xml import translate_drawings_and_charts
from .logging_utils import TranslationLogEntry
from .translators import RoutedTranslator

INVALID_SHEET_CHARS = r"[\\/*?:\[\]]"


@dataclass
class ProcessingResult:
    output_filename: str
    output_bytes: bytes
    logs: List[TranslationLogEntry]


def _safe_sheet_title(name: str, existing: set[str]) -> str:
    cleaned = re.sub(INVALID_SHEET_CHARS, "_", name).strip() or "Sheet"
    cleaned = cleaned[:31]
    base = cleaned
    i = 1
    while cleaned in existing:
        suffix = f"_{i}"
        cleaned = (base[: 31 - len(suffix)] + suffix)[:31]
        i += 1
    return cleaned


def _is_formula(cell: Cell) -> bool:
    if cell.data_type == "f":
        return True
    return isinstance(cell.value, str) and cell.value.startswith("=")


def _with_lang_suffix(name: str, target_lang: str) -> str:
    p = Path(name)
    return f"{p.stem}_{target_lang}{p.suffix}"


def process_excel_file(
    file_name: str,
    file_bytes: bytes,
    source_lang: str,
    target_lang: str,
    selected_engine: str,
) -> ProcessingResult:
    translator = RoutedTranslator(selected_engine=selected_engine)
    wb = load_workbook(io.BytesIO(file_bytes))
    logs: List[TranslationLogEntry] = []

    existing_titles: set[str] = set()
    for ws in wb.worksheets:
        original_title = ws.title
        try:
            translated_title, engine = translator.translate_with_engine(original_title, source_lang, target_lang)
            safe = _safe_sheet_title(translated_title, existing_titles)
            ws.title = safe
            existing_titles.add(safe)
            logs.append(
                TranslationLogEntry(
                    file_name=file_name,
                    sheet_name=safe,
                    object_id="sheet_title",
                    original_text=original_title,
                    translated_text=safe,
                    engine=engine,
                    status="ok",
                )
            )
        except Exception as exc:
            existing_titles.add(original_title)
            logs.append(
                TranslationLogEntry(
                    file_name=file_name,
                    sheet_name=original_title,
                    object_id="sheet_title",
                    original_text=original_title,
                    translated_text=original_title,
                    engine="none",
                    status="error",
                    error=str(exc),
                )
            )

    for ws in wb.worksheets:
        # Traverse deterministically by explicit row/column coordinates to ensure
        # full coverage across the worksheet grid boundaries.
        for row_idx in range(1, ws.max_row + 1):
            for col_idx in range(1, ws.max_column + 1):
                cell = ws.cell(row=row_idx, column=col_idx)

                if isinstance(cell.value, str) and not _is_formula(cell):
                    original = cell.value
                    try:
                        translated, engine = translator.translate_with_engine(original, source_lang, target_lang)
                        cell.value = translated
                        logs.append(
                            TranslationLogEntry(
                                file_name=file_name,
                                sheet_name=ws.title,
                                object_id=f"cell:{cell.coordinate}",
                                original_text=original,
                                translated_text=translated,
                                engine=engine,
                                status="ok",
                            )
                        )
                    except Exception as exc:
                        logs.append(
                            TranslationLogEntry(
                                file_name=file_name,
                                sheet_name=ws.title,
                                object_id=f"cell:{cell.coordinate}",
                                original_text=original,
                                translated_text=original,
                                engine="none",
                                status="error",
                                error=str(exc),
                            )
                        )

                if cell.comment and isinstance(cell.comment.text, str) and cell.comment.text.strip():
                    original_comment = cell.comment.text
                    try:
                        translated, engine = translator.translate_with_engine(original_comment, source_lang, target_lang)
                        cell.comment.text = translated
                        logs.append(
                            TranslationLogEntry(
                                file_name=file_name,
                                sheet_name=ws.title,
                                object_id=f"comment:{cell.coordinate}",
                                original_text=original_comment,
                                translated_text=translated,
                                engine=engine,
                                status="ok",
                            )
                        )
                    except Exception as exc:
                        logs.append(
                            TranslationLogEntry(
                                file_name=file_name,
                                sheet_name=ws.title,
                                object_id=f"comment:{cell.coordinate}",
                                original_text=original_comment,
                                translated_text=original_comment,
                                engine="none",
                                status="error",
                                error=str(exc),
                            )
                        )

    buf = io.BytesIO()
    wb.save(buf)
    wb.close()

    def translate_func(text: str, object_id: str) -> tuple[str, str]:
        translated, engine = translator.translate_with_engine(text, source_lang, target_lang)
        logs.append(
            TranslationLogEntry(
                file_name=file_name,
                sheet_name="<xml-layer>",
                object_id=object_id,
                original_text=text,
                translated_text=translated,
                engine=engine,
                status="ok",
            )
        )
        return translated, engine

    final_bytes, xml_logs = translate_drawings_and_charts(buf.getvalue(), translate_func)
    for item in xml_logs:
        if item.error:
            logs.append(
                TranslationLogEntry(
                    file_name=file_name,
                    sheet_name="<xml-layer>",
                    object_id=item.object_id,
                    original_text=item.original_text,
                    translated_text=item.translated_text,
                    engine="none",
                    status="error",
                    error=item.error,
                )
            )

    return ProcessingResult(output_filename=_with_lang_suffix(file_name, target_lang), output_bytes=final_bytes, logs=logs)
