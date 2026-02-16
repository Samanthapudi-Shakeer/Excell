from __future__ import annotations

import argparse
import io

from openpyxl import load_workbook


def validate(original_bytes: bytes, translated_bytes: bytes) -> None:
    wb_in = load_workbook(io.BytesIO(original_bytes))
    wb_out = load_workbook(io.BytesIO(translated_bytes))

    assert len(wb_in.sheetnames) == len(wb_out.sheetnames), "Sheet count changed"

    for ws_in, ws_out in zip(wb_in.worksheets, wb_out.worksheets):
        assert ws_in.max_row == ws_out.max_row, f"Row count mismatch in {ws_in.title}"
        assert ws_in.max_column == ws_out.max_column, f"Column count mismatch in {ws_in.title}"
        assert len(ws_in.merged_cells.ranges) == len(ws_out.merged_cells.ranges), f"Merged ranges mismatch in {ws_in.title}"

        for row in range(1, ws_in.max_row + 1):
            for col in range(1, ws_in.max_column + 1):
                c_in = ws_in.cell(row=row, column=col)
                c_out = ws_out.cell(row=row, column=col)
                if isinstance(c_in.value, str) and c_in.value.startswith("="):
                    assert c_in.value == c_out.value, f"Formula changed at {ws_in.title}!{c_in.coordinate}"
                if c_in.has_style:
                    assert c_in._style == c_out._style, f"Style changed at {ws_in.title}!{c_in.coordinate}"

    wb_in.close()
    wb_out.close()


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("original")
    parser.add_argument("translated")
    args = parser.parse_args()

    with open(args.original, "rb") as f:
        original = f.read()
    with open(args.translated, "rb") as f:
        translated = f.read()

    validate(original, translated)
    print("Validation passed")


if __name__ == "__main__":
    main()
