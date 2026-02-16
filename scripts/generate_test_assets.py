from __future__ import annotations

import io
from pathlib import Path

from openpyxl import Workbook
from openpyxl.chart import PieChart, Reference
from openpyxl.comments import Comment

ROOT = Path(__file__).resolve().parents[1]
ASSETS = ROOT / "tests" / "assets"


def create_sample(path: Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Résumé"
    ws["A1"] = "Hello 😀"
    ws["A2"] = "=SUM(B2:C2)"
    ws["B2"] = 1
    ws["C2"] = 2
    ws["D1"] = "Long text with special chars !@#$%^&*()"
    ws["A1"].comment = Comment("Commentaire", "bot")
    ws.merge_cells("A3:C3")
    ws["A3"] = "Merged block"

    ws2 = wb.create_sheet("Datos")
    ws2["A1"] = "Ventas"
    ws2["A2"] = "Q1"
    ws2["B2"] = 40
    ws2["A3"] = "Q2"
    ws2["B3"] = 60

    pie = PieChart()
    pie.title = "Distribution"
    labels = Reference(ws2, min_col=1, min_row=2, max_row=3)
    data = Reference(ws2, min_col=2, min_row=1, max_row=3)
    pie.add_data(data, titles_from_data=True)
    pie.set_categories(labels)
    ws2.add_chart(pie, "D4")

    wb.save(path)
    wb.close()


if __name__ == "__main__":
    ASSETS.mkdir(parents=True, exist_ok=True)
    create_sample(ASSETS / "sample_input.xlsx")
    print(f"Generated: {ASSETS / 'sample_input.xlsx'}")
